"""
后端端到端测试
覆盖所有交互流程：认证、积分、上传、管理员操作

运行方式：
  cd backend
  .venv/bin/pytest test_e2e.py -v
"""
import io
import pytest
from datetime import datetime, timedelta, timezone
from sqlalchemy import create_engine
from sqlalchemy.orm import sessionmaker
from sqlalchemy.pool import StaticPool
from unittest.mock import patch
from fastapi.testclient import TestClient
from openpyxl import load_workbook

# ─── 测试数据库配置（内存 SQLite，与生产库完全隔离）────────────────────────────

TEST_ENGINE = create_engine(
    "sqlite:///:memory:",
    connect_args={"check_same_thread": False},
    poolclass=StaticPool,
)
TestingSession = sessionmaker(autocommit=False, autoflush=False, bind=TEST_ENGINE)


def override_get_db():
    db = TestingSession()
    try:
        yield db
    finally:
        db.close()


# ─── 创建测试客户端（必须在 import main 之前 patch DB）────────────────────────

import config
config.RUN_TASKS_INLINE = True
import database
database.engine = TEST_ENGINE
database.SessionLocal = TestingSession
from database import Base, get_db, User, Video, Shot, VideoAnalysis, Credits, CreditTransaction, AnalysisTask
from main import app
import routers.analysis as analysis_router

app.dependency_overrides[get_db] = override_get_db

# 建表
Base.metadata.create_all(bind=TEST_ENGINE)

client = TestClient(app, raise_server_exceptions=True)


def clear_client_cookies():
    client.cookies.clear()


async def fake_run_analysis(_video_id, task_id, _user_id=None, _shot_indices=None):
    db = TestingSession()
    try:
        task = db.query(AnalysisTask).filter(AnalysisTask.id == task_id).first()
        if task:
            task.stage = "completed"
            task.done = task.total
            db.commit()
    finally:
        db.close()


ORIGINAL_RUN_ANALYSIS = analysis_router._run_analysis
analysis_router._run_analysis = fake_run_analysis


# ─── 工具函数 ───────────────────────────────────────────────────────────────────

def register(email, password="password123", display_name=""):
    return client.post("/api/auth/register", json={
        "email": email, "password": password, "display_name": display_name
    })


def login(email, password="password123"):
    return client.post("/api/auth/login", json={"email": email, "password": password})


def auth_headers(token: str):
    return {"Authorization": f"Bearer {token}"}


def get_db_session():
    return TestingSession()


def make_superuser(email: str):
    """将某用户设为超级管理员（直接操作测试 DB）"""
    db = get_db_session()
    try:
        user = db.query(User).filter(User.email == email).first()
        user.is_superuser = True
        db.commit()
    finally:
        db.close()


def set_credits(user_id: int, balance: int):
    """直接设置用户积分余额（用于测试 402 场景）"""
    db = get_db_session()
    try:
        credits = db.query(Credits).filter(Credits.user_id == user_id).first()
        credits.balance = balance
        db.commit()
    finally:
        db.close()


def create_video_with_shots(user_id: int, shot_count: int = 3) -> int:
    """在测试 DB 中直接创建视频+镜头记录，绕过真实上传"""
    db = get_db_session()
    try:
        video = Video(
            user_id=user_id,
            filename="test_video.mp4",
            filepath="/tmp/test_video.mp4",
            duration=30.0,
            fps=25.0,
            status="detected",
        )
        db.add(video)
        db.flush()
        for i in range(shot_count):
            db.add(Shot(
                video_id=video.id,
                index=i,
                start_time=float(i * 10),
                end_time=float(i * 10 + 9),
                duration=9.0,
            ))
        db.commit()
        return video.id
    finally:
        db.close()


# ═══════════════════════════════════════════════════════════════════════════════
# 测试用例
# ═══════════════════════════════════════════════════════════════════════════════

class TestHealth:
    def test_health_check(self):
        r = client.get("/health")
        assert r.status_code == 200
        assert r.json()["status"] == "ok"


class TestRegister:
    def test_register_success(self):
        r = register("new_user@example.com")
        assert r.status_code == 200
        data = r.json()
        assert "access_token" in data
        assert data["token_type"] == "bearer"
        assert isinstance(data["user_id"], int)

    def test_register_duplicate_email(self):
        register("dup@example.com")
        r = register("dup@example.com")
        assert r.status_code == 400
        assert "已注册" in r.json()["detail"]

    def test_register_short_password(self):
        r = register("short@example.com", password="123")
        assert r.status_code == 400
        assert "密码" in r.json()["detail"]

    def test_register_creates_credits(self):
        """注册后自动创建 100 积分"""
        r = register("credits_check@example.com")
        token = r.json()["access_token"]
        r2 = client.get("/api/credits/me", headers=auth_headers(token))
        assert r2.status_code == 200
        assert r2.json()["balance"] == 100

    def test_register_initial_grant_transaction(self):
        """注册后流水中应有 initial_grant 记录"""
        r = register("tx_check@example.com")
        token = r.json()["access_token"]
        r2 = client.get("/api/credits/me/transactions", headers=auth_headers(token))
        assert r2.status_code == 200
        data = r2.json()
        assert data["total"] >= 1
        assert any(t["reason"] == "initial_grant" for t in data["data"])


class TestLogin:
    @classmethod
    def setup_class(cls):
        register("login_test@example.com", "mypassword")

    def test_login_success(self):
        r = login("login_test@example.com", "mypassword")
        assert r.status_code == 200
        data = r.json()
        assert "access_token" in data
        assert "is_superuser" in data

    def test_login_wrong_password(self):
        r = login("login_test@example.com", "wrongpass")
        assert r.status_code == 401

    def test_login_nonexistent_user(self):
        r = login("nobody@example.com", "pass")
        assert r.status_code == 401

    def test_login_disabled_user(self):
        """禁用用户后登录应返回 403"""
        register("disabled@example.com", "pass123")
        # 直接在 DB 中禁用
        db = get_db_session()
        try:
            user = db.query(User).filter(User.email == "disabled@example.com").first()
            user.is_active = False
            db.commit()
        finally:
            db.close()
        r = login("disabled@example.com", "pass123")
        assert r.status_code == 403


class TestCurrentUser:
    @classmethod
    def setup_class(cls):
        r = register("me_test@example.com", display_name="测试用户")
        cls.token = r.json()["access_token"]

    def test_get_me(self):
        r = client.get("/api/auth/me", headers=auth_headers(self.token))
        assert r.status_code == 200
        data = r.json()
        assert data["email"] == "me_test@example.com"
        assert data["display_name"] == "测试用户"
        assert data["is_superuser"] == False

    def test_get_me_no_token(self):
        clear_client_cookies()
        r = client.get("/api/auth/me")
        assert r.status_code == 401

    def test_get_me_invalid_token(self):
        r = client.get("/api/auth/me", headers={"Authorization": "Bearer fake.token.here"})
        assert r.status_code == 401


class TestCredits:
    @classmethod
    def setup_class(cls):
        r = register("credits_user@example.com")
        cls.token = r.json()["access_token"]
        cls.user_id = r.json()["user_id"]

    def test_credits_balance(self):
        r = client.get("/api/credits/me", headers=auth_headers(self.token))
        assert r.status_code == 200
        assert r.json()["balance"] == 100

    def test_credits_transactions_list(self):
        r = client.get("/api/credits/me/transactions", headers=auth_headers(self.token))
        assert r.status_code == 200
        data = r.json()
        assert data["total"] >= 1
        assert data["balance"] == 100

    def test_credits_requires_auth(self):
        clear_client_cookies()
        r = client.get("/api/credits/me")
        assert r.status_code == 401


class TestUpload:
    @classmethod
    def setup_class(cls):
        r = register("uploader@example.com")
        cls.token = r.json()["access_token"]

    def _dummy_video(self, name="test.mp4"):
        """构造一个假的 mp4 文件（ffprobe 会失败，但端点会用默认值）"""
        return ("file", (name, io.BytesIO(b"fake video content"), "video/mp4"))

    def test_upload_as_user(self):
        r = client.post(
            "/api/upload",
            files=[self._dummy_video()],
            headers=auth_headers(self.token),
        )
        assert r.status_code == 200
        data = r.json()
        assert "video_id" in data
        assert data["filename"] == "test.mp4"

    def test_upload_requires_auth(self):
        clear_client_cookies()
        r = client.post("/api/upload", files=[self._dummy_video("anon.mp4")])
        assert r.status_code == 401

    def test_upload_invalid_format(self):
        r = client.post(
            "/api/upload",
            files=[("file", ("doc.txt", io.BytesIO(b"hello"), "text/plain"))],
            headers=auth_headers(self.token),
        )
        assert r.status_code == 400
        assert "不支持" in r.json()["detail"]

    def test_upload_rejects_video_over_duration_limit(self, monkeypatch):
        import routers.upload as upload_router

        monkeypatch.setattr(upload_router, "MAX_VIDEO_DURATION_SECONDS", 60)
        monkeypatch.setattr(upload_router, "get_video_meta", lambda _path: {
            "duration": 61.0,
            "fps": 25.0,
            "width": 1920,
            "height": 1080,
        })

        r = client.post(
            "/api/upload",
            files=[self._dummy_video("too_long.mp4")],
            headers=auth_headers(self.token),
        )

        assert r.status_code == 413
        assert "视频时长过长" in r.json()["detail"]


class TestVideoList:
    @classmethod
    def setup_class(cls):
        # 用户A 上传一个视频
        r_a = register("list_a@example.com")
        cls.token_a = r_a.json()["access_token"]
        cls.user_a_id = r_a.json()["user_id"]

        # 用户B 上传一个视频
        r_b = register("list_b@example.com")
        cls.token_b = r_b.json()["access_token"]

        client.post(
            "/api/upload",
            files=[("file", ("user_a.mp4", io.BytesIO(b"vid"), "video/mp4"))],
            headers=auth_headers(cls.token_a),
        )
        client.post(
            "/api/upload",
            files=[("file", ("user_b.mp4", io.BytesIO(b"vid"), "video/mp4"))],
            headers=auth_headers(cls.token_b),
        )

    def test_user_only_sees_own_videos(self):
        r = client.get("/api/videos", headers=auth_headers(self.token_a))
        assert r.status_code == 200
        filenames = [v["filename"] for v in r.json()]
        assert any("user_a" in f for f in filenames)
        assert not any("user_b" in f for f in filenames)

    def test_admin_workspace_only_sees_own_videos(self):
        r_admin = register("workspace_admin@example.com")
        admin_token = r_admin.json()["access_token"]
        make_superuser("workspace_admin@example.com")
        admin_token = login("workspace_admin@example.com").json()["access_token"]

        client.post(
            "/api/upload",
            files=[("file", ("admin_own.mp4", io.BytesIO(b"vid"), "video/mp4"))],
            headers=auth_headers(admin_token),
        )

        r = client.get("/api/videos", headers=auth_headers(admin_token))

        assert r.status_code == 200
        filenames = [v["filename"] for v in r.json()]
        assert any("admin_own" in f for f in filenames)
        assert not any("user_a" in f or "user_b" in f for f in filenames)

    def test_video_not_found(self):
        r = client.get("/api/videos/99999", headers=auth_headers(self.token_a))
        assert r.status_code == 404


class TestAdmin:
    @classmethod
    def setup_class(cls):
        # 普通用户
        r_normal = register("normal_for_admin@example.com")
        cls.normal_token = r_normal.json()["access_token"]

        # 管理员
        r_admin = register("admin@example.com")
        cls.admin_user_id = r_admin.json()["user_id"]
        make_superuser("admin@example.com")
        r_login = login("admin@example.com")
        cls.admin_token = r_login.json()["access_token"]

        # 受测用户
        r_target = register("target_user@example.com")
        cls.target_user_id = r_target.json()["user_id"]

    def test_normal_user_cannot_access_admin(self):
        r = client.get("/api/admin/users", headers=auth_headers(self.normal_token))
        assert r.status_code == 403

    def test_unauthenticated_cannot_access_admin(self):
        clear_client_cookies()
        r = client.get("/api/admin/users")
        assert r.status_code == 401

    def test_admin_list_users(self):
        r = client.get("/api/admin/users", headers=auth_headers(self.admin_token))
        assert r.status_code == 200
        data = r.json()
        assert "total" in data
        assert "data" in data
        assert data["total"] >= 2
        # 每个用户应包含必要字段
        for u in data["data"]:
            for field in ("id", "email", "credits", "video_count", "is_active"):
                assert field in u, f"缺少字段 {field}"

    def test_admin_list_users_keyword_filter(self):
        r = client.get(
            "/api/admin/users?keyword=target_user",
            headers=auth_headers(self.admin_token),
        )
        assert r.status_code == 200
        emails = [u["email"] for u in r.json()["data"]]
        assert any("target_user" in e for e in emails)

    def test_admin_get_user_detail(self):
        r = client.get(
            f"/api/admin/users/{self.target_user_id}",
            headers=auth_headers(self.admin_token),
        )
        assert r.status_code == 200
        data = r.json()
        assert data["id"] == self.target_user_id
        assert data["credits"] == 100

    def test_admin_get_nonexistent_user(self):
        r = client.get("/api/admin/users/99999", headers=auth_headers(self.admin_token))
        assert r.status_code == 404

    def test_admin_reset_credits(self):
        r = client.post(
            f"/api/admin/users/{self.target_user_id}/credits/reset",
            json={"balance": 500},
            headers=auth_headers(self.admin_token),
        )
        assert r.status_code == 200
        assert r.json()["new_balance"] == 500
        # 验证余额确实变了
        r2 = client.get(
            f"/api/admin/users/{self.target_user_id}",
            headers=auth_headers(self.admin_token),
        )
        assert r2.json()["credits"] == 500

    def test_admin_reset_credits_negative_rejected(self):
        r = client.post(
            f"/api/admin/users/{self.target_user_id}/credits/reset",
            json={"balance": -10},
            headers=auth_headers(self.admin_token),
        )
        assert r.status_code == 400

    def test_admin_reset_creates_transaction(self):
        """重置积分应产生 admin_reset 流水"""
        r = client.get(
            f"/api/admin/users/{self.target_user_id}/transactions",
            headers=auth_headers(self.admin_token),
        )
        assert r.status_code == 200
        reasons = [t["reason"] for t in r.json()["data"]]
        assert "admin_reset" in reasons

    def test_admin_disable_user(self):
        r = client.patch(
            f"/api/admin/users/{self.target_user_id}/status",
            headers=auth_headers(self.admin_token),
        )
        assert r.status_code == 200
        assert r.json()["is_active"] == False

    def test_admin_enable_user(self):
        """再次调用 toggle，恢复启用"""
        r = client.patch(
            f"/api/admin/users/{self.target_user_id}/status",
            headers=auth_headers(self.admin_token),
        )
        assert r.status_code == 200
        assert r.json()["is_active"] == True

    def test_admin_get_user_videos(self):
        r = client.get(
            f"/api/admin/users/{self.target_user_id}/videos",
            headers=auth_headers(self.admin_token),
        )
        assert r.status_code == 200
        data = r.json()
        assert "total" in data
        assert "data" in data


class TestCreditsAnalysis:
    """积分与分析流程：积分充足通过检查，不足返回 402"""

    @classmethod
    def setup_class(cls):
        r = register("analysis_user@example.com")
        cls.token = r.json()["access_token"]
        cls.user_id = r.json()["user_id"]

    def test_analyze_no_shots_returns_400(self):
        """视频没有镜头时，POST /analyze 应返回 400"""
        db = get_db_session()
        try:
            video = Video(
                user_id=self.user_id,
                filename="noshot.mp4",
                filepath="/tmp/noshot.mp4",
                duration=10.0,
                fps=25.0,
                status="detected",
            )
            db.add(video)
            db.commit()
            video_id = video.id
        finally:
            db.close()

        r = client.post(
            f"/api/analyze/{video_id}",
            headers=auth_headers(self.token),
        )
        assert r.status_code == 400
        assert "镜头检测" in r.json()["detail"]

    def test_analyze_sufficient_credits_passes_check(self):
        """余额 100，镜头数 3 → 通过积分检查，返回 200（后台任务启动）"""
        video_id = create_video_with_shots(self.user_id, shot_count=3)
        r = client.post(
            f"/api/analyze/{video_id}",
            headers=auth_headers(self.token),
        )
        # 200 = 积分检查通过，后台任务已启动
        assert r.status_code == 200
        assert r.json()["shot_count"] == 3

    def test_analyze_insufficient_credits_returns_402(self):
        """余额为 0，任何镜头数都应返回 402，且不能留下 analyzing 状态"""
        # 新建独立用户，确保余额干净
        r = register("broke_user@example.com")
        token = r.json()["access_token"]
        user_id = r.json()["user_id"]
        set_credits(user_id, 0)

        video_id = create_video_with_shots(user_id, shot_count=5)
        db = get_db_session()
        try:
            video = db.query(Video).filter(Video.id == video_id).first()
            video.status = "completed"
            video.error_msg = "旧错误不应被本次 402 改写"
            db.commit()
        finally:
            db.close()

        r = client.post(f"/api/analyze/{video_id}", headers=auth_headers(token))
        assert r.status_code == 402
        assert "积分不足" in r.json()["detail"]

        db = get_db_session()
        try:
            video = db.query(Video).filter(Video.id == video_id).first()
            assert video.status == "completed"
            assert video.current_task_id is None
            assert video.error_msg == "旧错误不应被本次 402 改写"
        finally:
            db.close()

    def test_analyze_nonexistent_video(self):
        r = client.post(
            "/api/analyze/99999",
            headers=auth_headers(self.token),
        )
        assert r.status_code == 404

    def test_analyze_requires_auth(self):
        clear_client_cookies()
        video_id = create_video_with_shots(self.user_id, shot_count=1)
        r = client.post(f"/api/analyze/{video_id}")
        assert r.status_code == 401


class TestShotAdjust:
    """手动调整镜头边界"""

    @classmethod
    def setup_class(cls):
        r = register("adjust_user@example.com")
        cls.token = r.json()["access_token"]
        cls.user_id = r.json()["user_id"]
        cls.video_id = create_video_with_shots(cls.user_id, shot_count=2)

    def test_adjust_shots(self):
        r = client.put(
            f"/api/shots/{self.video_id}/adjust",
            json={"shots": [
                {"start_time": 0.0, "end_time": 5.0},
                {"start_time": 5.0, "end_time": 12.0},
                {"start_time": 12.0, "end_time": 20.0},
            ]},
            headers=auth_headers(self.token),
        )
        assert r.status_code == 200
        assert r.json()["shot_count"] == 3

    def test_adjust_nonexistent_video(self):
        r = client.put(
            "/api/shots/99999/adjust",
            json={"shots": [{"start_time": 0.0, "end_time": 5.0}]},
            headers=auth_headers(self.token),
        )
        assert r.status_code == 404


class TestAuthorizationBoundaries:
    @classmethod
    def setup_class(cls):
        r_owner = register("owner@example.com")
        cls.owner_token = r_owner.json()["access_token"]
        cls.owner_id = r_owner.json()["user_id"]
        r_other = register("other@example.com")
        cls.other_token = r_other.json()["access_token"]
        cls.video_id = create_video_with_shots(cls.owner_id, shot_count=1)

    def test_other_user_cannot_get_results(self):
        r = client.get(f"/api/results/{self.video_id}", headers=auth_headers(self.other_token))
        assert r.status_code == 403

    def test_other_user_cannot_adjust_shots(self):
        r = client.put(
            f"/api/shots/{self.video_id}/adjust",
            json={"shots": [{"start_time": 0.0, "end_time": 1.0}]},
            headers=auth_headers(self.other_token),
        )
        assert r.status_code == 403

    def test_owner_can_get_results(self):
        r = client.get(f"/api/results/{self.video_id}", headers=auth_headers(self.owner_token))
        assert r.status_code == 200


class TestVideoPathFallback:
    def test_detect_uses_existing_shared_upload_when_db_has_stale_backend_path(self, tmp_path, monkeypatch):
        r = register("stale_path@example.com")
        token = r.json()["access_token"]
        user_id = r.json()["user_id"]

        stale_dir = tmp_path / "app" / "backend" / "uploads"
        shared_dir = tmp_path / "shared" / "uploads"
        stale_dir.mkdir(parents=True)
        shared_dir.mkdir(parents=True)
        shared_file = shared_dir / "same-name.mp4"
        shared_file.write_bytes(b"fake video")

        db = get_db_session()
        try:
            video = Video(
                user_id=user_id,
                filename="same-name.mp4",
                filepath=str(stale_dir / "same-name.mp4"),
                duration=10.0,
                fps=25.0,
                status="completed",
                error_msg="old error",
            )
            db.add(video)
            db.commit()
            video_id = video.id
        finally:
            db.close()

        from services.shot_detector import ShotBoundary

        seen = {}

        def fake_detect(path, threshold):
            seen["detect_path"] = path
            return [ShotBoundary(index=0, start_time=0.0, end_time=10.0, duration=10.0)]

        def fake_thumbs(path, shots, video_id):
            seen["thumb_path"] = path
            return [None]

        monkeypatch.setattr(config, "UPLOADS_DIR", shared_dir)
        monkeypatch.setattr("services.video_path.UPLOADS_DIR", shared_dir)
        monkeypatch.setattr(analysis_router, "detect_shots", fake_detect)
        monkeypatch.setattr("services.clip_extractor.extract_thumbnails_only", fake_thumbs)

        res = client.post(f"/api/detect/{video_id}", headers=auth_headers(token))

        assert res.status_code == 200
        assert seen["detect_path"] == str(shared_file)
        assert seen["thumb_path"] == str(shared_file)
        db = get_db_session()
        try:
            video = db.query(Video).filter(Video.id == video_id).first()
            assert video.filepath == str(shared_file)
            assert video.status == "detected"
            assert video.error_msg is None
        finally:
            db.close()


class TestClipExtractorTimestamps:
    def test_output_rate_rounds_fractional_source_rate(self):
        from fractions import Fraction
        from types import SimpleNamespace
        from services.clip_extractor import _output_rate

        assert _output_rate(SimpleNamespace(average_rate=Fraction(3097600, 129139))) == 24
        assert _output_rate(SimpleNamespace(average_rate=None)) == 25

    def test_audio_encoder_helper_exists_to_preserve_dialogue(self):
        from services.clip_extractor import _encode_audio_track

        assert callable(_encode_audio_track)


class TestAiAnalyzerClipBounds:
    def test_compute_extended_bounds_extends_first_shot_forward(self):
        from services.ai_analyzer import _compute_extended_bounds

        start, end = _compute_extended_bounds(0.0, 1.001, 15.133, 2.0)

        assert start == 0.0
        assert end >= 2.0

    def test_compute_extended_bounds_extends_middle_shot_backward(self):
        from services.ai_analyzer import _compute_extended_bounds

        start, end = _compute_extended_bounds(10.339, 11.506, 15.133, 2.0)

        assert start < 10.339
        assert end == 11.506
        assert end - start >= 2.0

    def test_build_merged_analysis_unit_first_shot_extends_forward(self):
        from types import SimpleNamespace
        from services.ai_analyzer import build_merged_analysis_unit

        shots = [
            SimpleNamespace(index=0, start_time=0.0, end_time=0.5),
            SimpleNamespace(index=1, start_time=0.5, end_time=1.6),
            SimpleNamespace(index=2, start_time=1.6, end_time=3.2),
        ]

        unit = build_merged_analysis_unit(shots, 0, safe_duration=3.0)

        assert unit["mode"] == "merged_context"
        assert unit["analysis_shot_indices"] == [0, 1, 2]
        assert unit["merged_duration"] >= 3.0
        assert unit["target_offset_start"] == 0.0

    def test_build_merged_analysis_unit_last_shot_extends_backward(self):
        from types import SimpleNamespace
        from services.ai_analyzer import build_merged_analysis_unit

        shots = [
            SimpleNamespace(index=0, start_time=0.0, end_time=1.2),
            SimpleNamespace(index=1, start_time=1.2, end_time=2.4),
            SimpleNamespace(index=2, start_time=2.4, end_time=3.1),
        ]

        unit = build_merged_analysis_unit(shots, 2, safe_duration=3.0)

        assert unit["mode"] == "merged_context"
        assert unit["analysis_shot_indices"] == [0, 1, 2]
        assert unit["merged_duration"] >= 3.0
        assert unit["target_offset_start"] == 2.4

    def test_build_merged_analysis_unit_middle_shot_extends_both_sides(self):
        from types import SimpleNamespace
        from services.ai_analyzer import build_merged_analysis_unit

        shots = [
            SimpleNamespace(index=0, start_time=0.0, end_time=1.3),
            SimpleNamespace(index=1, start_time=1.3, end_time=1.8),
            SimpleNamespace(index=2, start_time=1.8, end_time=3.1),
        ]

        unit = build_merged_analysis_unit(shots, 1, safe_duration=3.0)

        assert unit["analysis_shot_indices"] == [0, 1, 2]
        assert unit["target_offset_start"] == 1.3
        assert unit["target_offset_end"] == 1.8

    def test_transient_model_errors_are_retryable(self):
        from services.ai_analyzer import _is_transient_model_error

        assert _is_transient_model_error(Exception("Receive batching backend response failed!"))
        assert not _is_transient_model_error(Exception("The video file is too short."))


class TestAnalysisWorkerGuards:
    def test_failed_clip_is_not_sent_to_ai(self):
        db = get_db_session()
        try:
            r = register("failed_clip_user@example.com")
            user_id = r.json()["user_id"]
            video = Video(
                user_id=user_id,
                filename="failed_clip.mp4",
                filepath="/tmp/failed_clip.mp4",
                duration=10.0,
                fps=25.0,
                status="detected",
            )
            db.add(video)
            db.flush()
            shot = Shot(
                video_id=video.id,
                index=0,
                start_time=0.0,
                end_time=1.0,
                duration=1.0,
                clip_path=None,
            )
            db.add(shot)
            db.commit()
            video_id = video.id
        finally:
            db.close()

        from task_store import create_task
        create_task("missing_clip_task", video_id, user_id, 1, None)

        with (
            patch.object(analysis_router, "analyze_shot") as analyze_mock,
            patch.object(analysis_router, "extract_shot_clips", return_value=[(None, None)]),
        ):
            import asyncio
            asyncio.run(ORIGINAL_RUN_ANALYSIS(video_id, "missing_clip_task", user_id))
            analyze_mock.assert_not_called()

        db = get_db_session()
        try:
            stored = db.query(Shot).filter(Shot.video_id == video_id, Shot.index == 0).first()
            assert "切片失败" in stored.analysis["error"]
        finally:
            db.close()


class TestContextAnalyzer:
    def _shots(self, count=4, duration=4.0):
        from types import SimpleNamespace
        return [
            SimpleNamespace(index=i, start_time=i * duration, end_time=(i + 1) * duration, duration=duration)
            for i in range(count)
        ]

    def test_choose_strategy_whole_video_for_short_video(self):
        from services.context_analyzer import choose_analysis_strategy

        strategy = choose_analysis_strategy(30.0, 3)

        assert strategy.mode == "whole_video"

    def test_choose_strategy_selected_subset_uses_fallback(self):
        from services.context_analyzer import choose_analysis_strategy

        strategy = choose_analysis_strategy(30.0, 5, selected_count=2)

        assert strategy.mode == "shot_fallback"

    def test_choose_strategy_allows_whole_video_url_for_large_file(self, tmp_path, monkeypatch):
        import services.context_analyzer as context_analyzer

        video = tmp_path / "large.mp4"
        video.write_bytes(b"x" * 1024)
        monkeypatch.setattr(context_analyzer.app_config, "QWEN_VIDEO_INPUT_MODE", "auto")
        monkeypatch.setattr(context_analyzer.app_config, "PUBLIC_VIDEO_BASE_URL", "https://example.com/shotloom")
        monkeypatch.setattr(context_analyzer.app_config, "CONTEXT_BASE64_MAX_MB", 0.0001)

        strategy = context_analyzer.choose_analysis_strategy(30.0, 3, video_path=str(video))

        assert strategy.mode == "whole_video"

    def test_build_shot_chunks_respects_overlap(self):
        from services.context_analyzer import build_shot_chunks

        chunks = build_shot_chunks(self._shots(5, duration=2.0), max_duration=5.0, max_shots=3, overlap_shots=1)

        assert [[s.index for s in chunk] for chunk in chunks] == [[0, 1], [1, 2], [2, 3], [3, 4]]

    def test_normalize_context_indices_are_zero_based(self):
        from services.context_analyzer import _normalize_segment, _normalize_shot_result

        idx, analysis = _normalize_shot_result({"shot_index": 1, "what": "b"}, "whole_video")
        segment = _normalize_segment({"shot_indices": [0, 1], "summary": "s"}, "whole_video")

        assert idx == 1
        assert analysis["analysis_source"] == "whole_video"
        assert segment["shot_indices"] == [0, 1]


    def test_global_transcript_is_remapped_to_shot_boundaries(self, monkeypatch):
        from types import SimpleNamespace
        import services.context_analyzer as context_analyzer

        shots = [
            SimpleNamespace(index=0, start_time=0.0, end_time=3.7, duration=3.7),
            SimpleNamespace(index=1, start_time=3.7, end_time=7.2, duration=3.5),
            SimpleNamespace(index=2, start_time=7.2, end_time=12.7, duration=5.5),
            SimpleNamespace(index=3, start_time=12.7, end_time=15.0, duration=2.3),
        ]
        raw = {
            "global_transcript": [
                {"start_time": "0.0s", "end_time": "3.7s", "speaker": "旁白", "content": "第一句"},
                {"start_time": "3.7s", "end_time": "7.2s", "speaker": "旁白", "content": "第二句"},
                {"start_time": "7.2s", "end_time": "12.7s", "speaker": "旁白", "content": "第三句"},
                {"start_time": "12.7s", "end_time": "15.0s", "speaker": "旁白", "content": "第四句"},
            ],
            "shots": [
                {"shot_index": 0, "audio": {"dialogue": "第一句 第二句 第三句 第四句", "transcript_timestamps": "0.0s-15.0s"}},
                {"shot_index": 1, "audio": {"dialogue": "无", "transcript_timestamps": "无"}},
                {"shot_index": 2, "audio": {"dialogue": "无", "transcript_timestamps": "无"}},
                {"shot_index": 3, "audio": {"dialogue": "无", "transcript_timestamps": "无"}},
            ],
            "segments": [],
        }
        monkeypatch.setattr(context_analyzer, "_call_model_with_retries", lambda *_args, **_kwargs: raw)

        result = context_analyzer._call_context_model("video.mp4", shots, "whole_video")

        assert result["shots"][0]["audio"]["dialogue"] == "第一句"
        assert result["shots"][1]["audio"]["dialogue"] == "第二句"
        assert result["shots"][2]["audio"]["dialogue"] == "第三句"
        assert result["shots"][3]["audio"]["dialogue"] == "第四句"
        assert result["shots"][0]["audio"]["transcript_timestamps"] == "0.000s-3.700s"

    def test_chunk_prompt_uses_relative_times_with_original_context(self):
        from types import SimpleNamespace
        from services.context_analyzer import _context_prompt

        shot = SimpleNamespace(index=3, start_time=10.0, end_time=12.5, duration=2.5)

        prompt = _context_prompt([shot], "chunk_segment", start_time=10.0)

        assert "#3: 0.000s - 2.500s" in prompt
        assert "原视频时间 10.000s - 12.500s" in prompt
        assert "storyline" in prompt
        assert "不要把前一个/后一个镜头" in prompt
        assert "transcript_timestamps" in prompt

    def test_chunked_context_reports_each_chunk_as_it_finishes(self, tmp_path, monkeypatch):
        import services.context_analyzer as context_analyzer
        from types import SimpleNamespace

        shots = [
            SimpleNamespace(index=0, start_time=0.0, end_time=2.0, duration=2.0),
            SimpleNamespace(index=1, start_time=2.0, end_time=4.0, duration=2.0),
            SimpleNamespace(index=2, start_time=4.0, end_time=6.0, duration=2.0),
        ]
        monkeypatch.setattr(context_analyzer, "build_shot_chunks", lambda _shots: [[shots[0], shots[1]], [shots[2]]])
        monkeypatch.setattr(context_analyzer, "_extract_extended_clip", lambda *_args, **_kwargs: None)

        def fake_call(_path, chunk, source, start_time=0.0, chunk_index=None, video_id=None):
            return {
                "shots": {shot.index: {"what": f"shot-{shot.index}"} for shot in chunk},
                "segments": [{"shot_indices": [shot.index for shot in chunk], "summary": f"chunk-{chunk_index}"}],
            }

        monkeypatch.setattr(context_analyzer, "_call_context_model", fake_call)
        completed = []

        async def on_chunk(result, chunk, chunk_index, total_chunks):
            completed.append((chunk_index, total_chunks, [s.index for s in chunk], sorted(result["shots"].keys())))

        import asyncio
        result = asyncio.run(context_analyzer.analyze_chunked_context("source.mp4", shots, tmp_path, on_chunk_complete=on_chunk))

        assert completed == [(0, 2, [0, 1], [0, 1]), (1, 2, [2], [2])]
        assert sorted(result["shots"].keys()) == [0, 1, 2]


class TestContextAnalysisWorker:
    def _create_context_video(self, user_id: int, shot_count: int = 2) -> int:
        db = get_db_session()
        try:
            video = Video(
                user_id=user_id,
                filename="context.mp4",
                filepath="/tmp/context.mp4",
                duration=20.0,
                fps=25.0,
                status="detected",
            )
            db.add(video)
            db.flush()
            for i in range(shot_count):
                db.add(Shot(
                    video_id=video.id,
                    index=i,
                    start_time=float(i * 5),
                    end_time=float(i * 5 + 4),
                    duration=4.0,
                    clip_path=f"/tmp/context_{i}.mp4",
                ))
            db.commit()
            return video.id
        finally:
            db.close()

    def test_worker_saves_whole_video_context_and_segments(self):
        r = register("context_worker@example.com")
        user_id = r.json()["user_id"]
        video_id = self._create_context_video(user_id, shot_count=2)
        from task_store import create_task, get_task_progress
        create_task("context_task", video_id, user_id, 2, None)

        async def fake_context(_path, _shots, **_kwargs):
            return {
                "shots": {
                    0: {"what": "A", "analysis_source": "whole_video"},
                    1: {"what": "B", "analysis_source": "whole_video"},
                },
                "segments": [{"segment_index": 0, "shot_indices": [0, 1], "summary": "AB"}],
            }

        with (
            patch.object(analysis_router, "choose_analysis_strategy", return_value=type("S", (), {"mode": "whole_video", "reason": "test"})()),
            patch.object(analysis_router, "analyze_whole_video_context", side_effect=fake_context),
            patch.object(analysis_router, "analyze_shot") as analyze_mock,
            patch("pathlib.Path.exists", return_value=True),
        ):
            import asyncio
            asyncio.run(ORIGINAL_RUN_ANALYSIS(video_id, "context_task", user_id))
            analyze_mock.assert_not_called()

        db = get_db_session()
        try:
            shots = db.query(Shot).filter(Shot.video_id == video_id).order_by(Shot.index).all()
            va = db.query(VideoAnalysis).filter(VideoAnalysis.video_id == video_id).first()
            assert [shot.analysis["what"] for shot in shots] == ["A", "B"]
            assert va.segments_report["strategy"] == "whole_video"
            assert va.segments_report["segments"][0]["shot_indices"] == [0, 1]
            assert get_task_progress("context_task")["stage"] == "completed"
        finally:
            db.close()

    def test_worker_falls_back_for_missing_context_shot(self):
        r = register("context_fallback@example.com")
        user_id = r.json()["user_id"]
        video_id = self._create_context_video(user_id, shot_count=2)
        from task_store import create_task
        create_task("context_fallback_task", video_id, user_id, 2, None)

        async def fake_context(_path, _shots, **_kwargs):
            return {
                "shots": {0: {"what": "A", "analysis_source": "whole_video"}},
                "segments": [],
            }

        async def fake_analyze_shot(**_kwargs):
            return {"what": "fallback", "analysis_mode": "shot_clip"}

        with (
            patch.object(analysis_router, "choose_analysis_strategy", return_value=type("S", (), {"mode": "whole_video", "reason": "test"})()),
            patch.object(analysis_router, "analyze_whole_video_context", side_effect=fake_context),
            patch.object(analysis_router, "analyze_shot", side_effect=fake_analyze_shot) as analyze_mock,
            patch("pathlib.Path.exists", return_value=True),
        ):
            import asyncio
            asyncio.run(ORIGINAL_RUN_ANALYSIS(video_id, "context_fallback_task", user_id))
            assert analyze_mock.call_count == 1

        db = get_db_session()
        try:
            shots = db.query(Shot).filter(Shot.video_id == video_id).order_by(Shot.index).all()
            assert shots[0].analysis["what"] == "A"
            assert shots[1].analysis["what"] == "fallback"
        finally:
            db.close()


    def test_worker_updates_chunk_progress_and_saves_chunk_results_immediately(self):
        r = register("chunk_progress@example.com")
        user_id = r.json()["user_id"]
        video_id = self._create_context_video(user_id, shot_count=3)
        from task_store import create_task
        create_task("chunk_progress_task", video_id, user_id, 3, None)
        updates = []
        db_snapshots = []

        def fake_update_task(task_id, stage, done=None, total=None, msg=None):
            if task_id == "chunk_progress_task" and stage == "analyzing":
                updates.append((done, total, msg))

        async def fake_chunked(_path, chunk_shots, _temp_dir, video_id=None, on_chunk_complete=None):
            first = {"shots": {0: {"what": "A"}, 1: {"what": "B"}}, "segments": [{"shot_indices": [0, 1]}]}
            second = {"shots": {2: {"what": "C"}}, "segments": [{"shot_indices": [2]}]}
            await on_chunk_complete(first, chunk_shots[:2], 0, 2)
            db = get_db_session()
            try:
                db_snapshots.append([
                    s.analysis.get("what") if s.analysis else None
                    for s in db.query(Shot).filter(Shot.video_id == video_id).order_by(Shot.index).all()
                ])
            finally:
                db.close()
            await on_chunk_complete(second, chunk_shots[2:], 1, 2)
            return {"shots": {**first["shots"], **second["shots"]}, "segments": first["segments"] + second["segments"]}

        with (
            patch.object(analysis_router, "choose_analysis_strategy", return_value=type("S", (), {"mode": "chunk_segment", "reason": "test"})()),
            patch.object(analysis_router, "analyze_chunked_context", side_effect=fake_chunked),
            patch.object(analysis_router, "update_task", side_effect=fake_update_task),
            patch.object(analysis_router, "analyze_shot") as analyze_mock,
            patch("pathlib.Path.exists", return_value=True),
        ):
            import asyncio
            asyncio.run(ORIGINAL_RUN_ANALYSIS(video_id, "chunk_progress_task", user_id))
            analyze_mock.assert_not_called()

        assert db_snapshots == [["A", "B", None]]
        assert (1, 2, "分块上下文分析 1/2：已写入 2 个镜头") in updates
        assert (2, 2, "分块上下文分析 2/2：已写入 1 个镜头") in updates


class TestExportCompleteness:
    def _export_payload(self):
        video = {"id": 1, "filename": "demo.mp4", "duration": 12.5}
        shots = [{
            "index": 0,
            "start_time": 0.0,
            "end_time": 2.5,
            "duration": 2.5,
            "thumbnail_path": None,
            "analysis": {
                "shot_scale": "近景",
                "composition": "居中构图",
                "camera_movement": "固定",
                "lighting": "柔光",
                "color_tone": "暖色",
                "content_description": "角色抬头看向窗外",
                "on_screen_text": "字幕A",
                "time_evidence": "0.000s-2.500s",
                "dialogue": "你好",
                "audio": {
                    "dialogue": "你好，世界",
                    "speaker": "女声，普通话",
                    "sound_type": "人声+环境声",
                    "music": "轻柔钢琴",
                    "ambient_sound": "雨声",
                    "speaker_emotion": "平静",
                    "transcript_timestamps": "0.300s-1.200s",
                },
                "audiovisual_sync": "台词与表情同步",
                "audio_narrative_role": "交代人物状态",
                "audio_continuity": {
                    "continues_from_previous": False,
                    "continues_to_next": True,
                    "unfinished_dialogue": True,
                    "notes": "台词延续到下一镜头",
                },
                "action_continuity": {
                    "continues_from_previous": False,
                    "continues_to_next": True,
                    "notes": "抬头动作继续",
                },
                "what": "角色抬头",
                "how": "固定近景呈现表情",
                "why": "强调情绪变化",
                "narrative_level": {"scene": "室内", "event": "抬头", "information": "听到声音"},
                "emotional_function": "悬念",
                "narrative_decision": "延迟揭示",
                "rhythm_contribution": "中等节奏",
                "analysis_source": "chunk_segment",
                "custom_new_field": {"nested": "自定义字段保留"},
            },
        }]
        analysis = {
            "continuity": {
                "shot_scale_flow": "近景到特写",
                "movement_coherence": "动作顺接",
                "emotional_arc": "平静到紧张",
                "color_continuity": "暖色持续",
                "custom_audio_arc": "雨声贯穿",
            },
            "rhythm": {
                "avg_shot_duration": 2.5,
                "shortest_shot": 2.5,
                "longest_shot": 2.5,
                "plot_change_frequency": "稳定",
                "info_density_pattern": "逐步增加",
                "pacing_assessment": "节奏稳定",
                "tension_peaks": ["镜头1"],
            },
            "narrative_structure": {
                "detected_genre": "剧情",
                "three_act": "开端",
                "key_turning_points": ["镜头1"],
                "information_release_strategy": "延迟揭示",
            },
            "genre_patterns": {"structural_notes": "类型惯例", "deviation_notes": "无"},
            "custom_overall": {"score": "完整保留"},
        }
        segments = {
            "strategy": "chunk_segment",
            "reason": "test",
            "shot_count": 1,
            "segments": [{
                "segment_index": 0,
                "shot_indices": [0],
                "segment_type": "dialogue_continuity",
                "title": "开场段落",
                "summary": "角色听见声音",
                "merge_reason": "声音连续",
                "audio_continuity": "雨声持续",
                "action_continuity": "抬头动作延续",
                "editing_logic": "动作匹配",
                "emotional_arc": "悬念增强",
                "narrative_function": "建立情境",
                "custom_segment_field": {"detail": "段落自定义字段保留"},
            }],
        }
        return video, shots, analysis, segments

    def test_excel_export_includes_shot_segments_overall_and_complete_fields(self):
        from services.export_service import export_excel

        data = export_excel(*self._export_payload())
        wb = load_workbook(io.BytesIO(data))

        expected = {
            "导出说明",
            "镜头分析",
            "镜头完整JSON",
            "整体分析",
            "整体分析完整字段",
            "段落分析",
            "段落分析完整字段",
        }
        assert expected.issubset(set(wb.sheetnames))
        assert "时间证据" in [cell.value for cell in wb["镜头分析"][1]]
        assert "说话者/声线" in [cell.value for cell in wb["镜头分析"][1]]
        assert "台词时间戳" in [cell.value for cell in wb["镜头分析"][1]]
        assert "自定义字段保留" in wb["镜头完整JSON"]["B2"].value
        assert any(row[0] == "custom_overall.score" and row[1] == "完整保留" for row in wb["整体分析完整字段"].iter_rows(values_only=True))
        assert any(row[1] == "custom_segment_field.detail" and row[2] == "段落自定义字段保留" for row in wb["段落分析完整字段"].iter_rows(values_only=True))

    def test_pdf_html_export_includes_complete_json_sections(self):
        from services.export_service import export_pdf_html

        html = export_pdf_html(*self._export_payload())

        assert "完整镜头分析 JSON" in html
        assert "完整整体分析 JSON" in html
        assert "完整段落分析 JSON" in html
        assert "自定义字段保留" in html
        assert "段落自定义字段保留" in html
        assert "0.300s-1.200s" in html


class TestSignedVideoUrls:
    def test_signed_video_url_roundtrip_and_tamper_rejected(self):
        from services.signed_video_url import create_signed_video_token, verify_signed_video_token

        token = create_signed_video_token(42, expires_in=60, secret="test-secret")
        payload = verify_signed_video_token(token, secret="test-secret")

        assert payload["video_id"] == 42
        assert payload["scope"] == "ai_analysis"

        tampered = token[:-1] + ("a" if token[-1] != "a" else "b")
        with pytest.raises(ValueError):
            verify_signed_video_token(tampered, secret="test-secret")

    def test_public_signed_video_endpoint_streams_without_login(self, tmp_path):
        from services.signed_video_url import create_signed_video_token

        video_path = tmp_path / "public.mp4"
        video_path.write_bytes(b"fake video bytes")
        db = get_db_session()
        try:
            user = User(email="signed_video@example.com", hashed_password="x")
            db.add(user)
            db.flush()
            video = Video(
                user_id=user.id,
                filename="public.mp4",
                filepath=str(video_path),
                duration=1.0,
                fps=25.0,
                status="detected",
            )
            db.add(video)
            db.commit()
            video_id = video.id
        finally:
            db.close()

        token = create_signed_video_token(video_id, expires_in=60, secret=config.SIGNED_VIDEO_URL_SECRET)
        clear_client_cookies()
        r = client.get(f"/api/public/video/{token}")

        assert r.status_code == 200
        assert r.content == b"fake video bytes"
        assert r.headers["cache-control"] == "private, max-age=0, no-store"


class TestOmniVideoInput:
    def test_video_input_url_prefers_signed_public_url_for_large_file(self, tmp_path, monkeypatch):
        from services.ai_analyzer import _video_input_url

        video = tmp_path / "large.mp4"
        video.write_bytes(b"x" * 1024)
        monkeypatch.setattr(config, "QWEN_VIDEO_INPUT_MODE", "auto")
        monkeypatch.setattr(config, "PUBLIC_VIDEO_BASE_URL", "https://example.com/shotloom")
        monkeypatch.setattr(config, "CONTEXT_BASE64_MAX_MB", 0.0001)
        monkeypatch.setattr(config, "SIGNED_VIDEO_URL_SECRET", "test-secret")

        url = _video_input_url(str(video), video_id=7)

        assert url.startswith("https://example.com/shotloom/api/public/video/")
        assert not url.startswith("data:")

    def test_video_input_url_uses_base64_when_file_is_small(self, tmp_path, monkeypatch):
        from services.ai_analyzer import _video_input_url

        video = tmp_path / "small.mp4"
        video.write_bytes(b"small")
        monkeypatch.setattr(config, "QWEN_VIDEO_INPUT_MODE", "auto")
        monkeypatch.setattr(config, "PUBLIC_VIDEO_BASE_URL", "https://example.com/shotloom")
        monkeypatch.setattr(config, "CONTEXT_BASE64_MAX_MB", 10)

        url = _video_input_url(str(video), video_id=7)

        assert url.startswith("data:video/mp4;base64,")


    def test_default_model_is_qwen35_omni_plus(self):
        from pathlib import Path
        assert 'MODEL_NAME = os.getenv("MODEL_NAME", "qwen3.5-omni-plus")' in Path("config.py").read_text(encoding="utf-8")

    def test_omni_request_uses_stream_text_modalities(self, tmp_path, monkeypatch):
        import services.ai_analyzer as ai

        video = tmp_path / "clip.mp4"
        video.write_bytes(b"small")
        captured = {}

        class FakeCompletions:
            def create(self, **kwargs):
                captured.update(kwargs)
                delta = type("Delta", (), {"content": '{"ok": true}'})()
                choice = type("Choice", (), {"delta": delta})()
                return [type("Chunk", (), {"choices": [choice]})()]

        fake_client = type("Client", (), {
            "chat": type("Chat", (), {
                "completions": FakeCompletions()
            })()
        })()

        monkeypatch.setattr(ai, "_openai_client", fake_client)
        monkeypatch.setattr(ai, "MODEL_NAME", "qwen3.5-omni-plus")
        monkeypatch.setattr(ai, "QWEN_OMNI_OUTPUT_MODALITIES", ["text"])
        monkeypatch.setattr(config, "QWEN_VIDEO_INPUT_MODE", "base64")

        result = ai._call_omni_model(str(video), "prompt", video_id=99)

        assert result == {"ok": True}
        assert captured["model"] == "qwen3.5-omni-plus"
        assert captured["stream"] is True
        assert captured["stream_options"] == {"include_usage": True}
        assert captured["modalities"] == ["text"]
        content = captured["messages"][0]["content"]
        assert content[0]["type"] == "video_url"
        assert content[1] == {"type": "text", "text": "prompt"}

    def test_call_model_with_retries_passes_video_id_to_omni(self, tmp_path, monkeypatch):
        import services.ai_analyzer as ai

        video = tmp_path / "large.mp4"
        video.write_bytes(b"x")
        seen = {}

        def fake_call(path, prompt, video_id=None):
            seen["video_id"] = video_id
            return {"ok": True}

        monkeypatch.setattr(ai, "_is_omni_model", lambda: True)
        monkeypatch.setattr(ai, "_call_omni_model", fake_call)

        result = ai._call_model_with_retries(str(video), "prompt", video_id=123)

        assert result == {"ok": True}
        assert seen["video_id"] == 123


class TestProgress:
    def test_progress_not_found(self):
        r = client.get("/api/progress/nonexistent_task_id")
        assert r.status_code == 200
        # SSE 流：第一个事件应包含 not_found
        assert "not_found" in r.text

    def test_task_status_clears_stale_active_task(self):
        r = register("stale_task_user@example.com")
        token = r.json()["access_token"]
        user_id = r.json()["user_id"]
        video_id = create_video_with_shots(user_id, shot_count=3)
        task_id = "task_stale_refresh"
        stale_time = datetime.now(timezone.utc) - timedelta(minutes=31)

        db = get_db_session()
        try:
            video = db.query(Video).filter(Video.id == video_id).first()
            video.status = "analyzing"
            video.current_task_id = task_id
            db.add(AnalysisTask(
                id=task_id,
                video_id=video_id,
                user_id=user_id,
                stage="analyzing",
                done=0,
                total=3,
                updated_at=stale_time,
            ))
            db.commit()
        finally:
            db.close()

        r = client.get(f"/api/videos/{video_id}/task-status", headers=auth_headers(token))

        assert r.status_code == 200
        assert r.json()["has_active_task"] is False
        assert r.json()["task_id"] is None

        db = get_db_session()
        try:
            task = db.query(AnalysisTask).filter(AnalysisTask.id == task_id).first()
            video = db.query(Video).filter(Video.id == video_id).first()
            assert task.stage == "error"
            assert task.finished_at is not None
            assert "异常中断" in task.message
            assert video.status == "error"
            assert video.current_task_id is None
        finally:
            db.close()


# ─── 运行入口 ───────────────────────────────────────────────────────────────────

if __name__ == "__main__":
    import subprocess, sys
    subprocess.run([sys.executable, "-m", "pytest", __file__, "-v", "--tb=short"])
