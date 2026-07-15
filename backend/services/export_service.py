"""
导出服务：生成 Excel 和 PDF 报告
"""
import io
import base64
import html
import json
from pathlib import Path
from typing import Any, List

from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.drawing.image import Image as XLImage
from openpyxl.utils import get_column_letter


def _safe(val, default="—"):
    if val is None:
        return default
    if isinstance(val, list):
        return "；".join(str(v) for v in val)
    return str(val)


def _json_dump(value: Any) -> str:
    if value is None:
        return ""
    return json.dumps(value, ensure_ascii=False, indent=2)


def _flatten(value: Any, prefix: str = "") -> list[tuple[str, str]]:
    """Flatten nested analysis data so exports keep fields added by custom prompts."""
    if isinstance(value, dict):
        rows: list[tuple[str, str]] = []
        for key, child in value.items():
            child_prefix = f"{prefix}.{key}" if prefix else str(key)
            rows.extend(_flatten(child, child_prefix))
        return rows
    if isinstance(value, list):
        if all(not isinstance(item, (dict, list)) for item in value):
            return [(prefix, _safe(value))]
        rows = []
        for idx, child in enumerate(value):
            child_prefix = f"{prefix}[{idx}]"
            rows.extend(_flatten(child, child_prefix))
        return rows
    return [(prefix, _safe(value))]


def _style_header_row(ws, headers: list[str], fill, font, border):
    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.fill = fill
        cell.font = font
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = border


def _autosize_key_value_sheet(ws):
    ws.column_dimensions["A"].width = 34
    ws.column_dimensions["B"].width = 96
    for row in ws.iter_rows():
        for cell in row:
            cell.alignment = Alignment(vertical="top", wrap_text=True)


def _analysis_source_label(analysis: dict) -> str:
    source = analysis.get("analysis_source") or analysis.get("analysis_mode")
    mode = analysis.get("analysis_mode")
    if source == "whole_video" or mode == "whole_video_context":
        return "整片上下文"
    if source == "chunk_segment" or mode == "chunk_segment_context":
        return "分块上下文"
    if source == "merged_context" or mode == "merged_context":
        return "合并上下文"
    if source == "shot_clip" or mode == "shot_clip":
        return "单镜头"
    return _safe(source, "单镜头")


def _get_path(value: Any, path: str) -> Any:
    current = value
    for part in path.split("."):
        if not isinstance(current, dict):
            return None
        current = current.get(part)
    return current


def _schema_leaves(fields: list[dict[str, Any]], key_prefix: str = "", label_prefix: str = ""):
    for field in fields:
        key_path = f"{key_prefix}.{field['key']}" if key_prefix else field["key"]
        label_path = f"{label_prefix} / {field['label']}" if label_prefix else field["label"]
        children = field.get("fields") or []
        if children:
            yield from _schema_leaves(children, key_path, label_path)
        else:
            yield key_path, label_path, field


def _display_schema_value(value: Any) -> str | float | int | bool:
    if value is None or value == "":
        return "—"
    if isinstance(value, (dict, list)):
        return json.dumps(value, ensure_ascii=False)
    return value


def _export_excel_v2(video: dict, shots: List[dict], analysis: dict, segments: dict | None, schema: dict[str, Any]) -> bytes:
    wb = Workbook()
    header_fill = PatternFill("solid", fgColor="1F3864")
    header_font = Font(color="FFFFFF", bold=True, size=10)
    border = Border(left=Side(style="thin"), right=Side(style="thin"), top=Side(style="thin"), bottom=Side(style="thin"))
    scopes = schema.get("scopes") or {}

    manifest = wb.active
    manifest.title = "导出说明"
    manifest_rows = [
        ("视频文件", video.get("filename", "")),
        ("镜头数量", len(shots)),
        ("分析模板", schema.get("name", "")),
        ("Schema版本", schema.get("version", "")),
        ("视觉模型", video.get("vision_model", "qwen3.7-plus")),
        ("ASR模型", video.get("asr_model", "qwen3-asr-flash-filetrans")),
    ]
    for row_i, (key, value) in enumerate(manifest_rows, 1):
        manifest.cell(row=row_i, column=1, value=key)
        manifest.cell(row=row_i, column=2, value=value)
        for cell in manifest[row_i]:
            cell.border = border
            cell.alignment = Alignment(vertical="top", wrap_text=True)
    manifest.column_dimensions["A"].width = 24
    manifest.column_dimensions["B"].width = 90

    shot_leaves = list(_schema_leaves(scopes.get("shot") or []))
    ws = wb.create_sheet("镜头分析")
    base_headers = ["缩略图", "镜头#", "时长(s)", "开始", "结束"]
    _style_header_row(ws, base_headers + [label for _, label, _ in shot_leaves], header_fill, header_font, border)
    for col in range(1, len(base_headers) + len(shot_leaves) + 1):
        ws.column_dimensions[get_column_letter(col)].width = 12 if col <= 5 else 28
    for row_i, shot in enumerate(shots, 2):
        values = ["", int(shot.get("index", row_i - 2)) + 1, round(float(shot.get("duration") or 0), 3), shot.get("start_time"), shot.get("end_time")]
        result = shot.get("analysis") or {}
        values.extend(_display_schema_value(_get_path(result, path)) for path, _, _ in shot_leaves)
        for col_i, value in enumerate(values, 1):
            if col_i == 1:
                continue
            cell = ws.cell(row=row_i, column=col_i, value=value)
            cell.border = border
            cell.alignment = Alignment(vertical="top", wrap_text=True)
        thumb = shot.get("thumbnail_path")
        if thumb and Path(thumb).exists():
            try:
                img = XLImage(thumb)
                img.height, img.width = 56, 80
                ws.add_image(img, f"A{row_i}")
                ws.row_dimensions[row_i].height = 60
            except Exception:
                pass

    segment_leaves = list(_schema_leaves(scopes.get("segment") or []))
    ws_segments = wb.create_sheet("段落分析")
    _style_header_row(ws_segments, ["段落#", "镜头"] + [label for _, label, _ in segment_leaves], header_fill, header_font, border)
    for row_i, segment in enumerate((segments or {}).get("segments") or [], 2):
        values = [int(segment.get("segment_index", row_i - 2)) + 1, _safe([int(i) + 1 for i in segment.get("shot_indices", [])])]
        values.extend(_display_schema_value(_get_path(segment, path)) for path, _, _ in segment_leaves)
        for col_i, value in enumerate(values, 1):
            cell = ws_segments.cell(row=row_i, column=col_i, value=value)
            cell.border = border
            cell.alignment = Alignment(vertical="top", wrap_text=True)
            ws_segments.column_dimensions[get_column_letter(col_i)].width = 18 if col_i <= 2 else 32

    overall_leaves = list(_schema_leaves(scopes.get("overall") or []))
    ws_overall = wb.create_sheet("整体分析")
    _style_header_row(ws_overall, ["字段", "值"], header_fill, header_font, border)
    ws_overall.column_dimensions["A"].width = 34
    ws_overall.column_dimensions["B"].width = 96
    for row_i, (path, label, _) in enumerate(overall_leaves, 2):
        ws_overall.cell(row=row_i, column=1, value=label)
        ws_overall.cell(row=row_i, column=2, value=_display_schema_value(_get_path(analysis or {}, path)))
        for cell in ws_overall[row_i]:
            cell.border = border
            cell.alignment = Alignment(vertical="top", wrap_text=True)

    ws_json = wb.create_sheet("完整JSON")
    _style_header_row(ws_json, ["类型", "编号", "JSON"], header_fill, header_font, border)
    out_row = 2
    for shot in shots:
        ws_json.append(["镜头", int(shot.get("index", 0)) + 1, _json_dump(shot.get("analysis") or {})])
        out_row += 1
    for segment in (segments or {}).get("segments") or []:
        ws_json.append(["段落", int(segment.get("segment_index", 0)) + 1, _json_dump(segment)])
        out_row += 1
    ws_json.append(["整体", 1, _json_dump(analysis or {})])
    ws_json.column_dimensions["C"].width = 120
    for row in ws_json.iter_rows():
        for cell in row:
            cell.alignment = Alignment(vertical="top", wrap_text=True)

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def export_excel(video: dict, shots: List[dict], analysis: dict, segments: dict | None = None, schema: dict[str, Any] | None = None) -> bytes:
    if schema and int(schema.get("version") or 1) >= 2:
        return _export_excel_v2(video, shots, analysis, segments, schema)
    wb = Workbook()

    # ── Sheet 1: 镜头分析 ──────────────────────────────
    ws = wb.active
    ws.title = "镜头分析"

    header_fill = PatternFill("solid", fgColor="1F3864")
    alt_fill = PatternFill("solid", fgColor="EEF2FF")
    header_font = Font(color="FFFFFF", bold=True, size=10)
    border = Border(
        left=Side(style="thin"),
        right=Side(style="thin"),
        top=Side(style="thin"),
        bottom=Side(style="thin"),
    )

    headers = [
        "缩略图", "镜头#", "时长(s)", "开始", "结束",
        "景别", "运镜", "构图", "光影", "色调",
        "画面描述", "画面文字", "时间证据", "WHAT", "HOW", "WHY",
        "叙事-场景", "叙事-事件", "叙事-信息",
        "情绪功能", "叙事决策", "节奏贡献",
        "台词", "说话者/声线", "声音类型", "音乐", "环境声", "人声情绪", "台词时间戳",
        "声画关系", "声音叙事作用",
        "分析方式", "合并镜头", "合并段落分析",
        "声音承前", "声音启后", "台词未完", "声音连续说明",
        "动作承前", "动作启后", "动作连续说明", "错误",
    ]

    ws.row_dimensions[1].height = 20
    _style_header_row(ws, headers, header_fill, header_font, border)

    # 列宽
    col_widths = [12, 6, 8, 8, 8, 8, 8, 20, 18, 15,
                  36, 28, 24, 30, 30, 35, 20, 25, 25, 18,
                  30, 15, 32, 24, 18, 20, 22, 20, 20, 30,
                  14, 20, 40, 12, 12, 12, 30, 12, 12, 30, 28]
    for i, w in enumerate(col_widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w

    for row_i, shot in enumerate(shots, 2):
        a = shot.get("analysis") or {}
        nl = a.get("narrative_level") or {}
        audio = a.get("audio") or {}
        audio_continuity = a.get("audio_continuity") or {}
        action_continuity = a.get("action_continuity") or {}
        fill = alt_fill if row_i % 2 == 0 else PatternFill()

        ws.row_dimensions[row_i].height = 60

        values = [
            "",  # 缩略图占位
            shot.get("index", row_i - 1) + 1,
            round(shot.get("duration", 0), 1),
            f"{shot.get('start_time', 0):.1f}s",
            f"{shot.get('end_time', 0):.1f}s",
            _safe(a.get("shot_scale")),
            _safe(a.get("camera_movement")),
            _safe(a.get("composition")),
            _safe(a.get("lighting")),
            _safe(a.get("color_tone")),
            _safe(a.get("content_description")),
            _safe(a.get("on_screen_text")),
            _safe(a.get("time_evidence")),
            _safe(a.get("what")),
            _safe(a.get("how")),
            _safe(a.get("why")),
            _safe(nl.get("scene")),
            _safe(nl.get("event")),
            _safe(nl.get("information")),
            _safe(a.get("emotional_function")),
            _safe(a.get("narrative_decision")),
            _safe(a.get("rhythm_contribution")),
            _safe(audio.get("dialogue") or a.get("dialogue")),
            _safe(audio.get("speaker")),
            _safe(audio.get("sound_type")),
            _safe(audio.get("music")),
            _safe(audio.get("ambient_sound")),
            _safe(audio.get("speaker_emotion")),
            _safe(audio.get("transcript_timestamps")),
            _safe(a.get("audiovisual_sync")),
            _safe(a.get("audio_narrative_role")),
            _analysis_source_label(a),
            _safe([i + 1 for i in a.get("analysis_shot_indices", [])]) if a.get("analysis_shot_indices") else "—",
            _safe(a.get("merged_segment_analysis")),
            _safe(audio_continuity.get("continues_from_previous")),
            _safe(audio_continuity.get("continues_to_next")),
            _safe(audio_continuity.get("unfinished_dialogue")),
            _safe(audio_continuity.get("notes")),
            _safe(action_continuity.get("continues_from_previous")),
            _safe(action_continuity.get("continues_to_next")),
            _safe(action_continuity.get("notes")),
            _safe(a.get("error")),
        ]

        for col_i, val in enumerate(values, 1):
            if col_i == 1:
                continue  # 缩略图单独处理
            cell = ws.cell(row=row_i, column=col_i, value=val)
            cell.alignment = Alignment(vertical="center", wrap_text=True)
            cell.border = border
            if fill.fill_type:
                cell.fill = fill

        # 插入缩略图
        thumb = shot.get("thumbnail_path")
        if thumb and Path(thumb).exists():
            try:
                img = XLImage(thumb)
                img.height = 56
                img.width = 80
                ws.add_image(img, f"A{row_i}")
            except Exception:
                pass

    # ── Sheet 2: 镜头完整 JSON ─────────────────────────
    ws_json = wb.create_sheet("镜头完整JSON")
    _style_header_row(ws_json, ["镜头#", "完整分析 JSON"], header_fill, header_font, border)
    ws_json.column_dimensions["A"].width = 10
    ws_json.column_dimensions["B"].width = 120
    for row_i, shot in enumerate(shots, 2):
        ws_json.cell(row=row_i, column=1, value=shot.get("index", row_i - 2) + 1)
        ws_json.cell(row=row_i, column=2, value=_json_dump(shot.get("analysis") or {}))
        ws_json.row_dimensions[row_i].height = 120
        for cell in ws_json[row_i]:
            cell.alignment = Alignment(vertical="top", wrap_text=True)
            cell.border = border

    # ── Sheet 3: 整体分析 ──────────────────────────────
    ws2 = wb.create_sheet("整体分析")
    ws2.column_dimensions["A"].width = 25
    ws2.column_dimensions["B"].width = 80

    if analysis:
        rows = []
        c = analysis.get("continuity") or {}
        r = analysis.get("rhythm") or {}
        n = analysis.get("narrative_structure") or {}
        g = analysis.get("genre_patterns") or {}

        rows += [
            ("【连贯性】", ""),
            ("景别流动", _safe(c.get("shot_scale_flow"))),
            ("运镜衔接", _safe(c.get("movement_coherence"))),
            ("情绪弧线", _safe(c.get("emotional_arc"))),
            ("色调连续性", _safe(c.get("color_continuity"))),
            ("声音弧线", _safe(c.get("audio_arc"))),
            ("", ""),
            ("【节奏】", ""),
            ("平均镜头时长", f"{r.get('avg_shot_duration', '—')}s"),
            ("最短/最长镜头", f"{r.get('shortest_shot', '—')}s / {r.get('longest_shot', '—')}s"),
            ("剧情变化频率", _safe(r.get("plot_change_frequency"))),
            ("信息密度分布", _safe(r.get("info_density_pattern"))),
            ("节奏评估", _safe(r.get("pacing_assessment"))),
            ("张力高潮点", _safe(r.get("tension_peaks"))),
            ("", ""),
            ("【叙事结构】", ""),
            ("推测类型", _safe(n.get("detected_genre"))),
            ("三幕结构", _safe(n.get("three_act"))),
            ("关键转折点", _safe(n.get("key_turning_points"))),
            ("信息揭示策略", _safe(n.get("information_release_strategy"))),
            ("", ""),
            ("【类型规律】", ""),
            ("类型惯例体现", _safe(g.get("structural_notes"))),
            ("与惯例的偏差", _safe(g.get("deviation_notes"))),
        ]

        header_fill2 = PatternFill("solid", fgColor="2E4057")
        for r_i, (k, v) in enumerate(rows, 1):
            k_cell = ws2.cell(row=r_i, column=1, value=k)
            v_cell = ws2.cell(row=r_i, column=2, value=v)
            k_cell.alignment = Alignment(vertical="top", wrap_text=True)
            v_cell.alignment = Alignment(vertical="top", wrap_text=True)
            if k.startswith("【"):
                k_cell.font = Font(bold=True, color="FFFFFF")
                v_cell.font = Font(bold=True, color="FFFFFF")
                k_cell.fill = header_fill2
                v_cell.fill = header_fill2
            ws2.row_dimensions[r_i].height = max(15, min(120, len(str(v)) // 2))

        ws2_flat = wb.create_sheet("整体分析完整字段")
        _style_header_row(ws2_flat, ["字段", "值"], header_fill, header_font, border)
        _autosize_key_value_sheet(ws2_flat)
        for r_i, (key, value) in enumerate(_flatten(analysis), 2):
            ws2_flat.cell(row=r_i, column=1, value=key)
            ws2_flat.cell(row=r_i, column=2, value=value)
            ws2_flat.row_dimensions[r_i].height = max(18, min(140, len(str(value)) // 2))
        for row in ws2_flat.iter_rows():
            for cell in row:
                cell.border = border

    segment_rows = (segments or {}).get("segments") or []
    if segment_rows:
        ws3 = wb.create_sheet("段落分析")
        segment_headers = ["段落#", "镜头", "类型", "标题", "摘要", "合并原因", "声音连续", "动作连续", "剪辑逻辑", "情绪推进", "叙事功能"]
        widths = [8, 18, 18, 24, 50, 50, 36, 36, 40, 36, 40]
        for col, (header, width) in enumerate(zip(segment_headers, widths), 1):
            cell = ws3.cell(row=1, column=col, value=header)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
            cell.border = border
            ws3.column_dimensions[get_column_letter(col)].width = width
        for row_i, segment in enumerate(segment_rows, 2):
            indices = [int(i) + 1 for i in segment.get("shot_indices", [])]
            values = [
                segment.get("segment_index", row_i - 2) + 1,
                _safe(indices),
                _safe(segment.get("segment_type")),
                _safe(segment.get("title")),
                _safe(segment.get("summary")),
                _safe(segment.get("merge_reason")),
                _safe(segment.get("audio_continuity")),
                _safe(segment.get("action_continuity")),
                _safe(segment.get("editing_logic")),
                _safe(segment.get("emotional_arc")),
                _safe(segment.get("narrative_function")),
            ]
            ws3.row_dimensions[row_i].height = 72
            for col_i, val in enumerate(values, 1):
                cell = ws3.cell(row=row_i, column=col_i, value=val)
                cell.alignment = Alignment(vertical="top", wrap_text=True)
                cell.border = border
                if row_i % 2 == 0:
                    cell.fill = alt_fill

        ws3_flat = wb.create_sheet("段落分析完整字段")
        _style_header_row(ws3_flat, ["段落#", "字段", "值"], header_fill, header_font, border)
        ws3_flat.column_dimensions["A"].width = 10
        ws3_flat.column_dimensions["B"].width = 34
        ws3_flat.column_dimensions["C"].width = 96
        out_row = 2
        for segment in segment_rows:
            segment_no = int(segment.get("segment_index", out_row - 2)) + 1
            for key, value in _flatten(segment):
                ws3_flat.cell(row=out_row, column=1, value=segment_no)
                ws3_flat.cell(row=out_row, column=2, value=key)
                ws3_flat.cell(row=out_row, column=3, value=value)
                ws3_flat.row_dimensions[out_row].height = max(18, min(140, len(str(value)) // 2))
                for cell in ws3_flat[out_row]:
                    cell.alignment = Alignment(vertical="top", wrap_text=True)
                    cell.border = border
                out_row += 1

    manifest = wb.create_sheet("导出说明", 0)
    manifest.column_dimensions["A"].width = 24
    manifest.column_dimensions["B"].width = 90
    manifest_rows = [
        ("视频文件", video.get("filename", "")),
        ("镜头数量", len(shots)),
        ("包含内容", "镜头摘要、镜头完整 JSON、整体分析摘要、整体分析完整字段、段落分析摘要、段落分析完整字段"),
        ("说明", "完整字段页会保留自定义 prompt 新增字段，避免导出遗漏分析结果。"),
    ]
    for row_i, (key, value) in enumerate(manifest_rows, 1):
        manifest.cell(row=row_i, column=1, value=key)
        manifest.cell(row=row_i, column=2, value=value)
        for cell in manifest[row_i]:
            cell.alignment = Alignment(vertical="top", wrap_text=True)
            cell.border = border
        manifest.row_dimensions[row_i].height = 28

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def _export_pdf_html_v2(video: dict, shots: List[dict], analysis: dict, segments: dict | None, schema: dict[str, Any]) -> str:
    scopes = schema.get("scopes") or {}

    def esc(value: Any) -> str:
        return html.escape(str(_display_schema_value(value)))

    def render_fields(value: dict[str, Any], fields: list[dict[str, Any]]) -> str:
        blocks = []
        for field in fields:
            children = field.get("fields") or []
            if children:
                child_value = value.get(field["key"])
                child_html = render_fields(child_value if isinstance(child_value, dict) else {}, children)
                blocks.append(f'<section><h4>{esc(field["label"])}</h4>{child_html}</section>')
                continue
            item = value.get(field["key"])
            blocks.append(f'<div class="field"><b>{esc(field["label"])}</b><span>{esc(item)}</span></div>')
        return "".join(blocks)

    shot_cards = []
    for shot in shots:
        thumb = ""
        path = shot.get("thumbnail_path")
        if path and Path(path).exists():
            with open(path, "rb") as file:
                thumb = '<img src="data:image/jpeg;base64,' + base64.b64encode(file.read()).decode() + '">'
        shot_cards.append(
            f'<article class="card"><header>{thumb}<div><h3>镜头 #{int(shot.get("index", 0)) + 1}</h3>'
            f'<p>{float(shot.get("start_time", 0)):.3f}s → {float(shot.get("end_time", 0)):.3f}s · {float(shot.get("duration", 0)):.3f}s</p></div></header>'
            f'{render_fields(shot.get("analysis") or {}, scopes.get("shot") or [])}</article>'
        )
    segment_cards = []
    for segment in (segments or {}).get("segments") or []:
        indices = "、".join(f"#{int(index) + 1}" for index in segment.get("shot_indices", []))
        segment_cards.append(
            f'<article class="card"><h3>段落 #{int(segment.get("segment_index", 0)) + 1}</h3><p>镜头：{esc(indices)}</p>'
            f'{render_fields(segment, scopes.get("segment") or [])}</article>'
        )
    overall_html = render_fields(analysis or {}, scopes.get("overall") or [])
    return f"""<!doctype html><html lang="zh-CN"><head><meta charset="utf-8"><style>
@page{{size:A4;margin:15mm}}*{{box-sizing:border-box}}body{{font-family:'WenQuanYi Zen Hei','DejaVu Sans',sans-serif;color:#202034;font-size:11px}}
h1{{color:#1f3864;border-bottom:3px solid #4f46e5;padding-bottom:8px}}h2{{color:#312e81;margin-top:20px}}h3{{margin:4px 0;color:#312e81}}h4{{margin:10px 0 4px;color:#4338ca}}
.meta{{color:#6b7280}}.card{{border:1px solid #dfe3ee;border-left:4px solid #4f46e5;border-radius:7px;padding:10px;margin:10px 0;break-inside:avoid}}
header{{display:flex;gap:12px}}header img{{width:120px;height:68px;object-fit:cover;border-radius:4px}}section{{background:#f7f7fc;padding:7px;margin-top:7px;border-radius:5px}}
.field{{display:grid;grid-template-columns:120px 1fr;gap:8px;border-bottom:1px solid #ececf3;padding:4px 0;overflow-wrap:anywhere}}.field b{{color:#4b5563}}.field span{{white-space:pre-wrap}}
</style></head><body><h1>短视频分析报告 — {esc(video.get('filename', ''))}</h1>
<p class="meta">模板：{esc(schema.get('name', ''))} · Schema v{esc(schema.get('version', ''))} · 视觉 {esc(video.get('vision_model', 'qwen3.7-plus'))} · ASR {esc(video.get('asr_model', 'qwen3-asr-flash-filetrans'))}</p>
<h2>镜头分析</h2>{''.join(shot_cards)}<h2>段落分析</h2>{''.join(segment_cards) or '<p>暂无段落</p>'}
<h2>整体分析</h2><article class="card">{overall_html or '<p>暂无整体分析</p>'}</article></body></html>"""


def export_pdf_html(video: dict, shots: List[dict], analysis: dict, segments: dict | None = None, schema: dict[str, Any] | None = None) -> str:
    """生成 HTML 字符串，由 WeasyPrint 转 PDF"""

    if schema and int(schema.get("version") or 1) >= 2:
        return _export_pdf_html_v2(video, shots, analysis, segments, schema)

    def thumb_b64(path):
        if path and Path(path).exists():
            with open(path, "rb") as f:
                return "data:image/jpeg;base64," + base64.b64encode(f.read()).decode()
        return ""

    def esc(value):
        return html.escape(_safe(value))

    def json_pre(value):
        dumped = _json_dump(value)
        return html.escape(dumped) if dumped else "—"

    shot_cards = ""
    for shot in shots:
        a = shot.get("analysis") or {}
        nl = a.get("narrative_level") or {}
        audio = a.get("audio") or {}
        audio_continuity = a.get("audio_continuity") or {}
        action_continuity = a.get("action_continuity") or {}
        thumb = thumb_b64(shot.get("thumbnail_path"))
        img_tag = f'<img src="{thumb}" class="thumb">' if thumb else '<div class="no-thumb">无缩略图</div>'

        shot_cards += f"""
<div class="shot-card">
  <div class="shot-header">
    <div class="thumb-cell">{img_tag}</div>
    <div class="shot-meta">
      <h3>镜头 #{shot.get('index',0)+1}</h3>
      <p>时长：{shot.get('duration',0):.1f}s &nbsp;|&nbsp;
         {shot.get('start_time',0):.1f}s → {shot.get('end_time',0):.1f}s</p>
      <p><b>景别：</b>{esc(a.get('shot_scale'))} &nbsp;
         <b>运镜：</b>{esc(a.get('camera_movement'))}</p>
      <p><b>光影：</b>{esc(a.get('lighting'))}</p>
      <p><b>色调：</b>{esc(a.get('color_tone'))}</p>
      <p><b>画面文字：</b>{esc(a.get('on_screen_text'))}</p>
      <p><b>时间证据：</b>{esc(a.get('time_evidence'))}</p>
      {f"<p><b>分析方式：</b>{esc(_analysis_source_label(a))} &nbsp; <b>合并镜头：</b>{esc([i + 1 for i in a.get('analysis_shot_indices', [])])}</p>" if a.get('analysis_mode') == 'merged_context' else f"<p><b>分析方式：</b>{esc(_analysis_source_label(a))}</p>"}
    </div>
  </div>
  <div class="analysis-body">
    <p><b>画面描述：</b>{esc(a.get('content_description'))}</p>
    <div class="why-block">
      <div class="label">WHAT</div><div class="content">{esc(a.get('what'))}</div>
      <div class="label">HOW</div><div class="content">{esc(a.get('how'))}</div>
      <div class="label">WHY</div><div class="content why-text">{esc(a.get('why'))}</div>
    </div>
    <div class="narrative-block">
      <b>叙事层级</b>
      <p>场景：{esc(nl.get('scene'))}</p>
      <p>事件：{esc(nl.get('event'))}</p>
      <p>信息：{esc(nl.get('information'))}</p>
    </div>
    <div class="narrative-block">
      <b>声音</b>
      <p>台词：{esc(audio.get('dialogue') or a.get('dialogue'))}</p>
      <p>说话者/声线：{esc(audio.get('speaker'))}</p>
      <p>声音类型：{esc(audio.get('sound_type'))}</p>
      <p>音乐：{esc(audio.get('music'))}</p>
      <p>环境声：{esc(audio.get('ambient_sound'))}</p>
      <p>人声情绪：{esc(audio.get('speaker_emotion'))}</p>
      <p>台词时间戳：{esc(audio.get('transcript_timestamps'))}</p>
      <p>声画关系：{esc(a.get('audiovisual_sync'))}</p>
      <p>声音叙事：{esc(a.get('audio_narrative_role'))}</p>
    </div>
    <div class="narrative-block">
      <b>跨镜头连续性</b>
      <p>声音承前/启后：{esc(audio_continuity.get('continues_from_previous'))} / {esc(audio_continuity.get('continues_to_next'))}</p>
      <p>台词未完：{esc(audio_continuity.get('unfinished_dialogue'))}</p>
      <p>声音说明：{esc(audio_continuity.get('notes'))}</p>
      <p>动作承前/启后：{esc(action_continuity.get('continues_from_previous'))} / {esc(action_continuity.get('continues_to_next'))}</p>
      <p>动作说明：{esc(action_continuity.get('notes'))}</p>
    </div>
    <p><b>情绪功能：</b>{esc(a.get('emotional_function'))}</p>
    <p><b>叙事决策：</b>{esc(a.get('narrative_decision'))}</p>
    <p><b>节奏贡献：</b>{esc(a.get('rhythm_contribution'))}</p>
    <details open><summary>完整镜头分析 JSON</summary><pre>{json_pre(a)}</pre></details>
  </div>
</div>"""

    segments_html = ""
    segment_rows = (segments or {}).get("segments") or []
    if segment_rows:
        cards = ""
        for segment in segment_rows:
            cards += f"""
  <div class="segment-card">
    <h3>{esc(segment.get('title') or '段落')}</h3>
    <p><b>镜头：</b>{esc([int(i) + 1 for i in segment.get('shot_indices', [])])} &nbsp; <b>类型：</b>{esc(segment.get('segment_type'))}</p>
    <p><b>摘要：</b>{esc(segment.get('summary'))}</p>
    <p><b>合并原因：</b>{esc(segment.get('merge_reason'))}</p>
    <p><b>声音连续：</b>{esc(segment.get('audio_continuity'))}</p>
    <p><b>动作连续：</b>{esc(segment.get('action_continuity'))}</p>
    <p><b>剪辑逻辑：</b>{esc(segment.get('editing_logic'))}</p>
    <p><b>情绪推进：</b>{esc(segment.get('emotional_arc'))}</p>
    <p><b>叙事功能：</b>{esc(segment.get('narrative_function'))}</p>
    <details open><summary>完整段落分析 JSON</summary><pre>{json_pre(segment)}</pre></details>
  </div>"""
        segments_html = f"""
<div class="section">
  <h2>段落分析</h2>
  <p style="color:#6b7280;margin-bottom:10px">来源：{esc((segments or {}).get('strategy'))}；{esc((segments or {}).get('reason'))}</p>
  {cards}
</div>"""

    overall_html = ""
    if analysis:
        c = analysis.get("continuity") or {}
        r = analysis.get("rhythm") or {}
        n = analysis.get("narrative_structure") or {}
        overall_html = f"""
<div class="section">
  <h2>整体分析</h2>
  <h3>连贯性</h3>
  <p><b>景别流动：</b>{esc(c.get('shot_scale_flow'))}</p>
  <p><b>运镜衔接：</b>{esc(c.get('movement_coherence'))}</p>
  <p><b>情绪弧线：</b>{esc(c.get('emotional_arc'))}</p>
  <p><b>色调连续：</b>{esc(c.get('color_continuity'))}</p>
  <p><b>声音弧线：</b>{esc(c.get('audio_arc'))}</p>
  <h3>节奏</h3>
  <p><b>平均镜头时长：</b>{esc(r.get('avg_shot_duration','—'))}s</p>
  <p><b>节奏评估：</b>{esc(r.get('pacing_assessment'))}</p>
  <p><b>信息密度：</b>{esc(r.get('info_density_pattern'))}</p>
  <h3>叙事结构</h3>
  <p><b>推测类型：</b>{esc(n.get('detected_genre'))}</p>
  <p><b>三幕结构：</b>{esc(n.get('three_act'))}</p>
  <p><b>信息揭示策略：</b>{esc(n.get('information_release_strategy'))}</p>
  <details open><summary>完整整体分析 JSON</summary><pre>{json_pre(analysis)}</pre></details>
</div>"""

    return f"""<!DOCTYPE html>
<html lang="zh-CN">
<head>
<meta charset="UTF-8">
<style>
  @page {{ size: A4; margin: 16mm 14mm; }}
  * {{ box-sizing: border-box; margin: 0; padding: 0; }}
  body {{ font-family: 'WenQuanYi Zen Hei', 'DejaVu Sans', sans-serif; font-size: 11px;
          color: #1a1a2e; background: #f8f9ff; }}
  h1 {{ font-size: 20px; color: #1F3864; border-bottom: 3px solid #4f46e5;
        padding-bottom: 8px; margin-bottom: 16px; }}
  h2 {{ font-size: 16px; color: #312e81; margin: 20px 0 10px; }}
  h3 {{ font-size: 13px; color: #4338ca; margin: 12px 0 6px; }}
  .shot-card {{ background: white; border-radius: 8px; padding: 12px;
                margin-bottom: 16px; break-inside: avoid;
                border-left: 4px solid #4f46e5; box-shadow: 0 1px 4px rgba(0,0,0,.08); }}
  .shot-header {{ display: table; width: 100%; table-layout: fixed; margin-bottom: 10px; }}
  .thumb-cell {{ display: table-cell; width: 132px; padding-right: 12px; vertical-align: top; }}
  .shot-meta {{ display: table-cell; vertical-align: top; overflow-wrap: anywhere; word-break: break-word; }}
  .thumb {{ display: block; width: 120px; height: 68px; object-fit: cover; border-radius: 4px; }}
  .no-thumb {{ width: 120px; height: 68px; background: #e5e7eb; border-radius: 4px;
               display: block; line-height: 68px; text-align: center; color: #9ca3af; }}
  .shot-meta h3 {{ font-size: 14px; color: #1e1b4b; margin: 0 0 4px; }}
  .shot-meta p {{ margin: 2px 0; color: #374151; line-height: 1.55; }}
  .analysis-body {{ border-top: 1px solid #e5e7eb; padding-top: 8px; }}
  .why-block {{ background: #f5f3ff; border-radius: 6px; padding: 8px; margin: 8px 0; }}
  .label {{ font-weight: bold; color: #4f46e5; font-size: 10px;
            text-transform: uppercase; margin-top: 6px; }}
  .why-text {{ color: #1e1b4b; font-style: italic; }}
  .content {{ color: #374151; margin-bottom: 4px; overflow-wrap: anywhere; word-break: break-word; }}
  .narrative-block {{ background: #f0fdf4; border-radius: 6px; padding: 8px; margin: 6px 0; }}
  .narrative-block b {{ color: #166534; }}
  .section {{ background: white; border-radius: 8px; padding: 16px; margin-top: 20px; }}
  .segment-card {{ break-inside: avoid; border: 1px solid #e5e7eb; border-left: 4px solid #0ea5e9; border-radius: 8px; padding: 10px; margin: 10px 0; background: #f8fbff; }}
  details {{ margin-top: 8px; }}
  summary {{ color: #4338ca; font-weight: bold; margin-bottom: 4px; }}
  pre {{ white-space: pre-wrap; overflow-wrap: anywhere; word-break: break-word; font-size: 9px; line-height: 1.35; background: #f3f4f6; color: #111827; border-radius: 6px; padding: 8px; }}
  p {{ margin: 3px 0; line-height: 1.5; }}
</style>
</head>
<body>
<h1>拉片报告 — {video.get('filename','')}</h1>
<p style="color:#6b7280;margin-bottom:20px">
  时长：{video.get('duration',0):.1f}s &nbsp;|&nbsp; 共 {len(shots)} 个镜头
</p>
{shot_cards}
{segments_html}
{overall_html}
</body>
</html>"""
